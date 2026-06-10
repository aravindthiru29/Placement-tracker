import io
from datetime import date, datetime, timedelta
from flask import url_for
from models import db, User, LeetCodeProgress, AptitudeProgress, GitHubProgress, MonthlyProject
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

def compile_report_data(report_type, user_id=None, days=7):
    """
    Compiles report data for Weekly, Monthly, Member, or Group report views.
    - report_type: 'weekly', 'monthly', 'member', 'group'
    - user_id: Specific user if 'member' report
    - days: Time window (default 7 days for weekly, 30 for monthly)
    """
    today = date.today()
    start_date = today - timedelta(days=days)
    
    data = {
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'report_type': report_type.capitalize(),
        'period_start': start_date.strftime('%Y-%m-%d'),
        'period_end': today.strftime('%Y-%m-%d'),
        'summary': {},
        'details': []
    }
    
    if report_type == 'member' and user_id:
        user = User.query.get(user_id)
        if not user:
            return None
        data['member_name'] = user.name
        data['member_email'] = user.email
        
        # LeetCode History
        lc_logs = LeetCodeProgress.query.filter(
            LeetCodeProgress.user_id == user_id,
            LeetCodeProgress.date >= start_date
        ).order_by(LeetCodeProgress.date.desc()).all()
        
        # Aptitude History
        apt_logs = AptitudeProgress.query.filter(
            AptitudeProgress.user_id == user_id,
            AptitudeProgress.date >= start_date
        ).order_by(AptitudeProgress.date.desc()).all()
        
        # GitHub History
        gh_logs = GitHubProgress.query.filter(
            GitHubProgress.user_id == user_id,
            GitHubProgress.date >= start_date
        ).order_by(GitHubProgress.date.desc()).all()
        
        # Projects
        projects = MonthlyProject.query.filter_by(user_id=user_id).order_by(MonthlyProject.deadline.desc()).all()
        
        # LeetCode sums/stats
        lc_delta_easy = sum(l.easy_solved for l in lc_logs)
        lc_delta_med = sum(l.medium_solved for l in lc_logs)
        lc_delta_hard = sum(l.hard_solved for l in lc_logs)
        latest_lc = LeetCodeProgress.query.filter_by(user_id=user_id).order_by(LeetCodeProgress.date.desc()).first()
        
        # Aptitude sums/stats
        apt_q = sum(a.quant_questions + a.logical_questions + a.verbal_questions for a in apt_logs)
        latest_apt = AptitudeProgress.query.filter_by(user_id=user_id).order_by(AptitudeProgress.date.desc()).first()
        
        # GitHub sums/stats
        gh_commits = sum(g.commits for g in gh_logs)
        latest_gh = GitHubProgress.query.filter_by(user_id=user_id).order_by(GitHubProgress.date.desc()).first()
        
        data['summary'] = {
            'leetcode_easy_solved': lc_delta_easy,
            'leetcode_medium_solved': lc_delta_med,
            'leetcode_hard_solved': lc_delta_hard,
            'leetcode_total_solved': lc_delta_easy + lc_delta_med + lc_delta_hard,
            'leetcode_current_total': latest_lc.total_solved if latest_lc else 0,
            'leetcode_streak': latest_lc.streak if latest_lc else 0,
            
            'aptitude_questions_solved': apt_q,
            'aptitude_accuracy': latest_apt.accuracy_percentage if latest_apt else 0.0,
            'aptitude_mock_score': latest_apt.mock_test_score if latest_apt else 0.0,
            
            'github_commits': gh_commits,
            'github_pull_requests': sum(g.pull_requests for g in gh_logs),
            'github_features_completed': sum(g.features_completed for g in gh_logs),
            'github_repos_created': sum(g.repositories_created for g in gh_logs),
            
            'projects_count': len(projects),
            'projects_completed': sum(1 for p in projects if p.status == 'Completed')
        }
        
        data['details'] = {
            'leetcode': [{'date': l.date.strftime('%Y-%m-%d'), 'easy': l.easy_solved, 'medium': l.medium_solved, 'hard': l.hard_solved, 'total': l.total_solved, 'streak': l.streak} for l in lc_logs],
            'aptitude': [{'date': a.date.strftime('%Y-%m-%d'), 'quant': a.quant_questions, 'logical': a.logical_questions, 'verbal': a.verbal_questions, 'accuracy': a.accuracy_percentage, 'mock_score': a.mock_test_score} for a in apt_logs],
            'github': [{'date': g.date.strftime('%Y-%m-%d'), 'commits': g.commits, 'repos': g.repositories_created, 'features': g.features_completed, 'prs': g.pull_requests} for g in gh_logs],
            'projects': [{'name': p.project_name, 'deadline': p.deadline.strftime('%Y-%m-%d'), 'progress': p.completion_percentage, 'status': p.status, 'github': p.github_link} for p in projects]
        }
        
    else:
        # Group summary
        members = User.query.filter_by(role='member').all()
        data['total_members'] = len(members)
        
        # Calculate group aggregates
        total_lc = 0
        total_apt_q = 0
        total_commits = 0
        total_projects = 0
        completed_projects = 0
        
        member_summaries = []
        for m in members:
            # Latest scores
            latest_lc = LeetCodeProgress.query.filter_by(user_id=m.id).order_by(LeetCodeProgress.date.desc()).first()
            latest_apt = AptitudeProgress.query.filter_by(user_id=m.id).order_by(AptitudeProgress.date.desc()).first()
            latest_gh = GitHubProgress.query.filter_by(user_id=m.id).order_by(GitHubProgress.date.desc()).first()
            projs = MonthlyProject.query.filter_by(user_id=m.id).all()
            
            lc_solved = latest_lc.total_solved if latest_lc else 0
            apt_solved = AptitudeProgress.query.filter_by(user_id=m.id).order_by(AptitudeProgress.date.desc()).first()
            apt_q_val = (latest_apt.quant_questions + latest_apt.logical_questions + latest_apt.verbal_questions) if latest_apt else 0
            commits_val = latest_gh.commits if latest_gh else 0
            
            total_lc += lc_solved
            total_apt_q += apt_q_val
            total_commits += commits_val
            total_projects += len(projs)
            completed_projects += sum(1 for p in projs if p.status == 'Completed')
            
            member_summaries.append({
                'id': m.id,
                'name': m.name,
                'email': m.email,
                'leetcode_solved': lc_solved,
                'leetcode_streak': latest_lc.streak if latest_lc else 0,
                'aptitude_solved': apt_q_val,
                'aptitude_accuracy': latest_apt.accuracy_percentage if latest_apt else 0.0,
                'github_commits': commits_val,
                'projects_created': len(projs),
                'projects_completed': sum(1 for p in projs if p.status == 'Completed')
            })
            
        data['summary'] = {
            'group_total_leetcode': total_lc,
            'group_avg_leetcode': round(total_lc / len(members), 1) if members else 0,
            'group_total_aptitude_questions': total_apt_q,
            'group_avg_aptitude_questions': round(total_apt_q / len(members), 1) if members else 0,
            'group_total_commits': total_commits,
            'group_avg_commits': round(total_commits / len(members), 1) if members else 0,
            'group_total_projects': total_projects,
            'group_project_completion_rate': round((completed_projects / total_projects) * 100, 1) if total_projects else 0.0
        }
        
        data['details'] = member_summaries
        
    return data

def export_report_to_excel(report_data):
    """
    Exports the report data into a beautiful, styled Excel spreadsheet using openpyxl.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # Styles
    title_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    section_font = Font(name='Segoe UI', size=12, bold=True, color='1F4E79')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Segoe UI', size=11)
    bold_data_font = Font(name='Segoe UI', size=11, bold=True)
    
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    zebra_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    accent_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # -------------------------------------------------------------
    # SHEET 1: OVERVIEW / SUMMARY
    # -------------------------------------------------------------
    ws_summary = wb.create_sheet(title='Summary Overview')
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Header Banner
    ws_summary.merge_cells('A1:D1')
    ws_summary['A1'] = f"Placement Tracker - {report_data['report_type']} Report"
    ws_summary['A1'].font = title_font
    ws_summary['A1'].fill = title_fill
    ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.row_dimensions[1].height = 40
    
    # Metadata info
    ws_summary['A3'] = "Generated At:"
    ws_summary['A3'].font = bold_data_font
    ws_summary['B3'] = report_data['generated_at']
    ws_summary['B3'].font = data_font
    
    ws_summary['A4'] = "Reporting Period:"
    ws_summary['A4'].font = bold_data_font
    ws_summary['B4'] = f"{report_data['period_start']} to {report_data['period_end']}"
    ws_summary['B4'].font = data_font
    
    if 'member_name' in report_data:
        ws_summary['A5'] = "Member Name:"
        ws_summary['A5'].font = bold_data_font
        ws_summary['B5'] = report_data['member_name']
        ws_summary['B5'].font = data_font
        ws_summary['A6'] = "Member Email:"
        ws_summary['A6'].font = bold_data_font
        ws_summary['B6'] = report_data['member_email']
        ws_summary['B6'].font = data_font
        
    # Stats table
    start_row = 8
    ws_summary.cell(row=start_row, column=1, value="Metric Category").font = section_font
    ws_summary.cell(row=start_row, column=2, value="Key Performance Indicator").font = section_font
    ws_summary.cell(row=start_row, column=3, value="Value").font = section_font
    
    ws_summary.row_dimensions[start_row].height = 24
    
    kpis = []
    if 'member_name' in report_data:
        kpis = [
            ("LeetCode", "Easy Problems Solved (Period)", report_data['summary']['leetcode_easy_solved']),
            ("LeetCode", "Medium Problems Solved (Period)", report_data['summary']['leetcode_medium_solved']),
            ("LeetCode", "Hard Problems Solved (Period)", report_data['summary']['leetcode_hard_solved']),
            ("LeetCode", "Total Problems Solved (Period)", report_data['summary']['leetcode_total_solved']),
            ("LeetCode", "Current Streak (Days)", report_data['summary']['leetcode_streak']),
            ("Aptitude", "Questions Solved (Period)", report_data['summary']['aptitude_questions_solved']),
            ("Aptitude", "Latest Average Accuracy", f"{report_data['summary']['aptitude_accuracy']}%"),
            ("Aptitude", "Latest Mock Score", f"{report_data['summary']['aptitude_mock_score']}%"),
            ("GitHub", "Commits Pushed (Period)", report_data['summary']['github_commits']),
            ("GitHub", "Pull Requests Raised (Period)", report_data['summary']['github_pull_requests']),
            ("GitHub", "Features Completed (Period)", report_data['summary']['github_features_completed']),
            ("GitHub", "Repositories Created (Period)", report_data['summary']['github_repos_created']),
            ("Projects", "Total Monthly Projects", report_data['summary']['projects_count']),
            ("Projects", "Completed Projects", report_data['summary']['projects_completed']),
        ]
    else:
        kpis = [
            ("Group Overview", "Total Managed Members", report_data['total_members']),
            ("LeetCode", "Group Total Solved", report_data['summary']['group_total_leetcode']),
            ("LeetCode", "Group Average Solved per Member", report_data['summary']['group_avg_leetcode']),
            ("Aptitude", "Group Total Questions Solved", report_data['summary']['group_total_aptitude_questions']),
            ("Aptitude", "Group Average Questions per Member", report_data['summary']['group_avg_aptitude_questions']),
            ("GitHub", "Group Total Commits", report_data['summary']['group_total_commits']),
            ("GitHub", "Group Average Commits per Member", report_data['summary']['group_avg_commits']),
            ("Projects", "Group Total Project Registrations", report_data['summary']['group_total_projects']),
            ("Projects", "Group Project Completion Rate", f"{report_data['summary']['group_project_completion_rate']}%"),
        ]
        
    curr_row = start_row + 1
    for cat, kpi, val in kpis:
        c1 = ws_summary.cell(row=curr_row, column=1, value=cat)
        c2 = ws_summary.cell(row=curr_row, column=2, value=kpi)
        c3 = ws_summary.cell(row=curr_row, column=3, value=val)
        
        for c in [c1, c2, c3]:
            c.font = data_font
            c.border = thin_border
            if curr_row % 2 == 0:
                c.fill = zebra_fill
        c3.alignment = Alignment(horizontal='right')
        curr_row += 1
        
    # Auto-adjust column widths for Summary sheet
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # -------------------------------------------------------------
    # SHEET 2: DETAILED DATA (Details or Member Comparison)
    # -------------------------------------------------------------
    if 'member_name' in report_data:
        # Create separate tabs for LeetCode, Aptitude, GitHub logs
        sections = [
            ('LeetCode Logs', ['Date', 'Easy Solved', 'Medium Solved', 'Hard Solved', 'Total Solved', 'Streak'], 'leetcode'),
            ('Aptitude Logs', ['Date', 'Quant Solved', 'Logical Solved', 'Verbal Solved', 'Accuracy %', 'Mock Score %'], 'aptitude'),
            ('GitHub Logs', ['Date', 'Commits', 'Repos Created', 'Features Completed', 'Pull Requests'], 'github'),
            ('Projects List', ['Project Name', 'Deadline', 'Completion %', 'Status', 'GitHub URL'], 'projects')
        ]
        
        for title, headers, key in sections:
            ws = wb.create_sheet(title=title)
            ws.views.sheetView[0].showGridLines = True
            
            # Header Row
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            ws.row_dimensions[1].height = 26
            
            # Write data rows
            items = report_data['details'][key]
            for row_idx, item in enumerate(items, 2):
                ws.row_dimensions[row_idx].height = 20
                vals = []
                if key == 'leetcode':
                    vals = [item['date'], item['easy'], item['medium'], item['hard'], item['total'], item['streak']]
                elif key == 'aptitude':
                    vals = [item['date'], item['quant'], item['logical'], item['verbal'], f"{item['accuracy']}%", f"{item['mock_score']}%"]
                elif key == 'github':
                    vals = [item['date'], item['commits'], item['repos'], item['features'], item['prs']]
                elif key == 'projects':
                    vals = [item['name'], item['deadline'], f"{item['progress']}%", item['status'], item['github']]
                    
                for col_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = data_font
                    cell.border = thin_border
                    if row_idx % 2 == 1:
                        cell.fill = zebra_fill
                    
                    # Alignments
                    if col_idx == 1 or (key == 'projects' and col_idx in [2, 4]):
                        cell.alignment = Alignment(horizontal='center')
                    elif col_idx > 1 and key != 'projects':
                        cell.alignment = Alignment(horizontal='right')
            
            # Auto-adjust column width
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
    else:
        # Group Member Comparison sheet
        ws_members = wb.create_sheet(title='Member Rankings')
        ws_members.views.sheetView[0].showGridLines = True
        
        headers = ['Member Name', 'Email', 'LeetCode Solved', 'LeetCode Streak', 'Aptitude Questions', 'Aptitude Accuracy %', 'GitHub Commits', 'Projects Created', 'Projects Completed']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_members.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        ws_members.row_dimensions[1].height = 26
        
        for row_idx, member in enumerate(report_data['details'], 2):
            ws_members.row_dimensions[row_idx].height = 20
            vals = [
                member['name'],
                member['email'],
                member['leetcode_solved'],
                member['leetcode_streak'],
                member['aptitude_solved'],
                f"{member['aptitude_accuracy']}%",
                member['github_commits'],
                member['projects_created'],
                member['projects_completed']
            ]
            
            for col_idx, val in enumerate(vals, 1):
                cell = ws_members.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                if row_idx % 2 == 1:
                    cell.fill = zebra_fill
                    
                # Alignments
                if col_idx in [1, 2]:
                    cell.alignment = Alignment(horizontal='left')
                elif col_idx in [6]:
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='center')
                    
        # Column width adjustments
        for col in ws_members.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_members.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    # Save workbook to memory buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

def export_report_to_pdf(report_data):
    """
    Generates a beautifully styled, professional PDF report using reportlab.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40, leftMargin=40,
        topMargin=40, bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor('#1F4E79'),
        spaceAfter=15,
        alignment=0 # Left
    )
    
    meta_style = ParagraphStyle(
        'MetaText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#555555'),
        spaceAfter=5
    )
    
    h2_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor('#2F5597'),
        spaceBefore=15,
        spaceAfter=10,
        keepWithNext=True
    )
    
    body_style = ParagraphStyle(
        'TableBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11
    )
    
    header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white
    )
    
    story = []
    
    # Header Section
    story.append(Paragraph(f"Placement Tracker System", title_style))
    story.append(Paragraph(f"<b>Report Type:</b> {report_data['report_type']} Performance Audit", meta_style))
    story.append(Paragraph(f"<b>Reporting Period:</b> {report_data['period_start']} to {report_data['period_end']}", meta_style))
    story.append(Paragraph(f"<b>Generated At:</b> {report_data['generated_at']}", meta_style))
    if 'member_name' in report_data:
        story.append(Paragraph(f"<b>Audited Member:</b> {report_data['member_name']} ({report_data['member_email']})", meta_style))
    story.append(Spacer(1, 15))
    
    # Summary Section
    story.append(Paragraph("Performance Summary Metrics", h2_style))
    
    # Compile summary table
    summary_data = [[Paragraph("Metric Category", header_style), Paragraph("Key Performance Indicator", header_style), Paragraph("Value", header_style)]]
    
    kpis = []
    if 'member_name' in report_data:
        kpis = [
            ("LeetCode", "Easy Solved (This Period)", str(report_data['summary']['leetcode_easy_solved'])),
            ("LeetCode", "Medium Solved (This Period)", str(report_data['summary']['leetcode_medium_solved'])),
            ("LeetCode", "Hard Solved (This Period)", str(report_data['summary']['leetcode_hard_solved'])),
            ("LeetCode", "Total Solved (This Period)", str(report_data['summary']['leetcode_total_solved'])),
            ("LeetCode", "Current Streak", f"{report_data['summary']['leetcode_streak']} Days"),
            ("Aptitude", "Questions Solved (This Period)", str(report_data['summary']['aptitude_questions_solved'])),
            ("Aptitude", "Latest Avg Accuracy", f"{report_data['summary']['aptitude_accuracy']}%"),
            ("Aptitude", "Latest Mock Test Score", f"{report_data['summary']['aptitude_mock_score']}%"),
            ("GitHub", "Commits Pushed (This Period)", str(report_data['summary']['github_commits'])),
            ("GitHub", "PRs Raised / Features Solved", f"{report_data['summary']['github_pull_requests']} / {report_data['summary']['github_features_completed']}"),
            ("Monthly Projects", "Projects Completed / Active", f"{report_data['summary']['projects_completed']} / {report_data['summary']['projects_count']}"),
        ]
    else:
        kpis = [
            ("Group Summary", "Total Monitored Members", str(report_data['total_members'])),
            ("LeetCode", "Group Total Solved Problems", str(report_data['summary']['group_total_leetcode'])),
            ("LeetCode", "Group Average Solved / Member", str(report_data['summary']['group_avg_leetcode'])),
            ("Aptitude", "Group Total Aptitude Solved", str(report_data['summary']['group_total_aptitude_questions'])),
            ("Aptitude", "Group Average Aptitude Solved / Member", str(report_data['summary']['group_avg_aptitude_questions'])),
            ("GitHub", "Group Total Commits", str(report_data['summary']['group_total_commits'])),
            ("GitHub", "Group Average Commits / Member", str(report_data['summary']['group_avg_commits'])),
            ("Projects", "Group Total Projects Logged", str(report_data['summary']['group_total_projects'])),
            ("Projects", "Group Project Completion Rate", f"{report_data['summary']['group_project_completion_rate']}%"),
        ]
        
    for cat, kpi, val in kpis:
        summary_data.append([
            Paragraph(cat, body_style),
            Paragraph(kpi, body_style),
            Paragraph(val, body_style)
        ])
        
    t_summary = Table(summary_data, colWidths=[1.5*inch, 3.5*inch, 2.0*inch])
    t_summary.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2F5597')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D9D9D9')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F2F2F2')]),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,1), (-1,-1), 4),
        ('BOTTOMPADDING', (0,1), (-1,-1), 4),
    ]))
    story.append(t_summary)
    
    story.append(Spacer(1, 15))
    
    # Detail Section
    if 'member_name' in report_data:
        story.append(Paragraph("Monthly Projects Status", h2_style))
        proj_headers = [Paragraph("Project Name", header_style), Paragraph("Deadline", header_style), Paragraph("Progress", header_style), Paragraph("Status", header_style)]
        proj_table_data = [proj_headers]
        
        for p in report_data['details']['projects']:
            proj_table_data.append([
                Paragraph(p['name'], body_style),
                Paragraph(p['deadline'], body_style),
                Paragraph(f"{p['progress']}%", body_style),
                Paragraph(p['status'], body_style)
            ])
            
        t_proj = Table(proj_table_data, colWidths=[3.2*inch, 1.2*inch, 1.1*inch, 1.5*inch])
        t_proj.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2F5597')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D9D9D9')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F2F2F2')]),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(t_proj)
        
        story.append(PageBreak())
        
        # LeetCode activity table
        story.append(Paragraph("LeetCode Logging History", h2_style))
        lc_headers = [Paragraph("Date", header_style), Paragraph("Easy", header_style), Paragraph("Medium", header_style), Paragraph("Hard", header_style), Paragraph("Total Current", header_style), Paragraph("Streak", header_style)]
        lc_table_data = [lc_headers]
        
        for l in report_data['details']['leetcode'][:15]: # Show latest 15 records
            lc_table_data.append([
                Paragraph(l['date'], body_style),
                Paragraph(str(l['easy']), body_style),
                Paragraph(str(l['medium']), body_style),
                Paragraph(str(l['hard']), body_style),
                Paragraph(str(l['total']), body_style),
                Paragraph(f"{l['streak']} Days", body_style)
            ])
        
        t_lc = Table(lc_table_data, colWidths=[1.5*inch, 1.0*inch, 1.0*inch, 1.0*inch, 1.3*inch, 1.2*inch])
        t_lc.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2F5597')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D9D9D9')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F2F2F2')]),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(t_lc)
        
    else:
        # Group ranks
        story.append(Paragraph("Member Rankings & Aggregate Metrics", h2_style))
        rank_headers = [
            Paragraph("Rank", header_style),
            Paragraph("Name", header_style),
            Paragraph("LeetCode", header_style),
            Paragraph("Streak", header_style),
            Paragraph("Aptitude Qs", header_style),
            Paragraph("Accuracy", header_style),
            Paragraph("Commits", header_style),
            Paragraph("Projects Completed", header_style)
        ]
        rank_table_data = [rank_headers]
        
        # Sort details by leetcode solved or just rank them
        sorted_members = sorted(report_data['details'], key=lambda x: x['leetcode_solved'], reverse=True)
        for idx, m in enumerate(sorted_members, 1):
            rank_table_data.append([
                Paragraph(str(idx), body_style),
                Paragraph(m['name'], body_style),
                Paragraph(str(m['leetcode_solved']), body_style),
                Paragraph(f"{m['leetcode_streak']}d", body_style),
                Paragraph(str(m['aptitude_solved']), body_style),
                Paragraph(f"{m['aptitude_accuracy']}%", body_style),
                Paragraph(str(m['github_commits']), body_style),
                Paragraph(f"{m['projects_completed']}/{m['projects_created']}", body_style)
            ])
            
        t_ranks = Table(rank_table_data, colWidths=[0.5*inch, 1.8*inch, 0.9*inch, 0.7*inch, 0.9*inch, 0.9*inch, 0.8*inch, 1.5*inch])
        t_ranks.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2F5597')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D9D9D9')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F2F2F2')]),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(t_ranks)
        
    doc.build(story)
    buffer.seek(0)
    return buffer
